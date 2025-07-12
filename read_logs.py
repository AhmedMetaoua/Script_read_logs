import re
import sys
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import pymysql

# Configuration de la base de données
DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "azerty",
    "database": "scheduler_test"
}

def ajouter_dans_bdd(date_traitement, heure_execution, donnees):
    try:
        connection = pymysql.connect(**DB_CONFIG)
        cursor = connection.cursor()

        # Création de la table si elle n'existe pas
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS Suivi_Btj (
                id INT AUTO_INCREMENT PRIMARY KEY,
                date_traitement DATE,
                heure_execution TIME,
                cptrs_releves INT,
                cptrs_non_releves INT,
                cptrs_ni_ai INT,
                abonnes_factures INT,
                abonnes_non_factures INT,
                factures_ordinaire INT,
                factures_simple INT,
                factures_adm INT,
                somme_restitutions_adm INT
            )
        """)

        # Insertion des données
        insert_query = """
            INSERT INTO Suivi_Btj (
                date_traitement, heure_execution,
                cptrs_releves, cptrs_non_releves, cptrs_ni_ai,
                abonnes_factures, abonnes_non_factures,
                factures_ordinaire, factures_simple,
                factures_adm, somme_restitutions_adm
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        cursor.execute(insert_query, (date_traitement, heure_execution, *donnees))
        connection.commit()
        print("✅ Données ajoutées à la base de données.")
    except pymysql.MySQLError as e:
        print(f"❌ Erreur base de données: {e}")
    finally:
        cursor.close()
        connection.close()

# Vérification des arguments
if len(sys.argv) != 2:
    print("❌ Usage: python script.py <fichier_releve.txt>")
    sys.exit(1)

# Récupérer le nom de fichier depuis la ligne de commande
fichier_texte = sys.argv[1]
fichier_excel = "Suivi Btj 2025.xlsx"

# Lire le fichier texte
try:
    with open(fichier_texte, 'r', encoding='utf-8') as f:
        contenu = f.read()
except FileNotFoundError:
    print(f"❌ Fichier introuvable: {fichier_texte}")
    sys.exit(1)

# Extraire la date de traitement
date_match = re.search(r"to_date\('(\d{2}/\d{2}/\d{4})'", contenu)
date_traitement = datetime.strptime(date_match.group(1), "%d/%m/%Y").date() if date_match else None

# Extraire l'heure depuis la première ligne
heure_match = re.search(r'Production on \w+ \w+ \d+ (\d{2}:\d{2}:\d{2}) \d{4}', contenu)
heure_execution = heure_match.group(1) if heure_match else "00:00:00"

# Extraire les données
valeurs = {
    "Cptrs Releves": re.search(r"Cptrs a Releves\s*-+\s*(\d+)", contenu),
    "Cptrs Non Releves": re.search(r"Cptrs Non Releves\s*-+\s*(\d+)", contenu),
    "Cptrs NI = AI": re.search(r"Cptrs NI egale AI\s*-+\s*(\d+)", contenu),
    "Abonnes Factures": re.search(r"Abonnes Factures\s*-+\s*(\d+)", contenu),
    "Abonnes Non Factures": re.search(r"Abon\.Non Factures\s*-+\s*(\d+)", contenu),
    "Factures Ordinaire": re.search(r"Nbre Factures Fichier Ordinaire\s*-+\s*(\d+)", contenu),
    "Factures Simple": re.search(r"Nbre Factures Fichier Simple\s*-+\s*(\d+)", contenu),
    "Factures ADM": re.search(r"Factures ADM de la mensuel de prelevement\s*-+\s*(\d+)", contenu),
    "Somme Restitutions + ADM": re.search(r"Somme Restitutions et Adm\s*-+\s*(\d+)", contenu)
}

donnees = [int(m.group(1)) if m else 0 for m in valeurs.values()]
ligne = [str(date_traitement), heure_execution] + donnees

# Ajouter ou créer le fichier Excel
if os.path.exists(fichier_excel):
    wb = load_workbook(fichier_excel)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Heure", *valeurs.keys()])

ws.append(ligne)
wb.save(fichier_excel)
print(f"✅ Données extraites de {fichier_texte} et ajoutées à '{fichier_excel}'.")

# Ajouter à la base de données
ajouter_dans_bdd(date_traitement, heure_execution, donnees)

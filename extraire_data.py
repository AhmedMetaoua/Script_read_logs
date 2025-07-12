import re
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os

# Fichiers
fichier_texte = "releve.txt"
fichier_excel = "Suivi Btj 2025.xlsx"

# Lire le fichier texte
with open(fichier_texte, 'r', encoding='utf-8') as f:
    contenu = f.read()

# Extraire la date de traitement
date_match = re.search(r"to_date\('(\d{2}/\d{2}/\d{4})'", contenu)
date_traitement = datetime.strptime(date_match.group(1), "%d/%m/%Y").date() if date_match else None

# Extraire l'heure depuis la première ligne du fichier
heure_match = re.search(r'Production on \w+ \w+ \d+ (\d{2}:\d{2}:\d{2}) \d{4}', contenu)
heure_execution = heure_match.group(1) if heure_match else ""

# Extraire les différentes données numériques
valeurs = {
    "Cptrs Releves": re.search(r"Cptrs a Releves\s*-+\s*(\d+)", contenu),
    "Cptrs Non Releves": re.search(r"Cptrs Non Releves\s*-+\s*(\d+)", contenu),
    "Cptrs NI = AI": re.search(r"Cptrs NI egale AI\s*-+\s*(\d+)", contenu),
    "Abonnes Factures": re.search(r"Abonnes Factures\s*-+\s*(\d+)", contenu),
    "Abonnes Non Factures": re.search(r"Abon\.Non Factures\s*-+\s*(\d+)", contenu),
    "Factures Ordinaire": re.search(r"Nbre Factures Fichier Ordinaire\s*-+\s*(\d+)", contenu),
    "Factures Simple": re.search(r"Nbre Factures Fichier Simple\s*-+\s*(\d+)", contenu),
    "Factures ADM": re.search(r"Factures ADM de la mensuel de prelevement\s*-+\s*(\d+)", contenu),
    "Somme Restitutions + ADM": re.search(r"Somme Restitutions et Adm\s*-+\s*(\d+)", contenu)
}

# Convertir en valeurs numériques
donnees = [int(m.group(1)) if m else 0 for m in valeurs.values()]

# Créer la ligne à ajouter
ligne = [str(date_traitement), heure_execution] + donnees

# Ajouter au fichier Excel
if os.path.exists(fichier_excel):
    wb = load_workbook(fichier_excel)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    # Entêtes
    ws.append(["Date", "Heure", *valeurs.keys()])

# Ajouter la ligne à la fin
ws.append(ligne)

# Sauvegarder
wb.save(fichier_excel)
print("✅ Données ajoutées dans le fichier Excel avec succès.")
